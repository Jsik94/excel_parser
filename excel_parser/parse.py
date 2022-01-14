import csv
import re

import pandas as pd
import xlrd
from openpyxl import load_workbook
import os


class ExcelExtractor:

    def __init__(self,dir):
        self.origin= dir
        self.dir_ = os.path.dirname(os.path.realpath(dir)) + '{}dataset{}'.format(os.path.sep, os.path.sep)
        print('loc', self.dir_)


    def getFileList(self):
        self.file_list = os.listdir(self.dir_)
        print('getFileList:',self.file_list)




    def extract(self):
        print(self.dir_ + self.file_list[0])
        self.data_list =[]
        self.data_colum=["회사명","url","담당자 번호","주소","연봉","인센","근무지","담당 업무","요건"]
        for list_ in self.file_list:
            print('Extract target file :', list_)
            # res =pd.read_excel(path+os.path.sep+list_)
            wb = xlrd.open_workbook(self.dir_+list_)
            sheet = wb.sheet_by_index(0)
            sten = str(sheet.cell(14,2).value);
            res = re.search(r'\d{2,}',sten)
            data_record =[]

            data_record.append(sheet.cell(2,1).value)
            data_record.append(sheet.cell(2,4).value)
            data_record.append(sheet.cell(4,2).value)
            data_record.append(sheet.cell(6,1).value)
            data_record.append("협상" if res is None else "{} {}".format(res.group(),"만원"))
            data_record.append(sheet.cell(15,2).value)
            data_record.append(sheet.cell(17,1).value)
            data_record.append(str(sheet.cell(9,1).value).replace("\n"," "))
            data_record.append(sheet.cell(10,2).value)
            data_record.append(str(sheet.cell(12,2).value).replace("\n"," "))
            dict_ = dict()
            # print(dict(zip(data_colum,data_record)))

            ##파일명이 필요한 경우만
            # data_total = dict()
            # data_total["filename"]=list_
            # data_total["record"]=dict(zip(data_colum,data_record))
            # print(data_total)
            # self.data_list.append(data_total)
            self.data_list.append(dict(zip(self.data_colum,data_record)))


    def saveCSV(self):
        dist_path = os.path.dirname(os.path.realpath(self.origin))+"{}data_dist{}".format(os.path.sep,os.path.sep)
        # print(self.data_list)
        print(dist_path+"result.csv")
        print(self.origin)
        print(self.data_list)
        with open(dist_path+"result.csv".format(os.path.sep),'w',encoding='utf8',newline='') as f:
            dw = csv.DictWriter(f,fieldnames=self.data_colum)
            dw.writeheader()
            for list_ in self.data_list:
                print(list_)
                dw.writerow(list_)


        print('완료되었습니다')